"""
报表导出为 Excel (.xlsx)。

用法：
  python scripts/export_xlsx.py --analysis <分析类型> --output <文件路径> [筛选参数...]

分析类型与 analyze_report.py 一致，共 8 种。
"""

import argparse
import os
import sys
from collections import defaultdict

from _import_base import *  # noqa: F401,F403

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from report_utils import make_admin_headers, output_result
from config import BASE_URL

# 复用 analyze_report 的数据拉取和分析逻辑
from analyze_report import (
    ANALYSIS_TYPES,
    fetch_all_tasks,
    fetch_all_resources,
    fetch_accel_cards,
    fetch_resource_specs,
    fetch_clusters,
    fetch_nodes,
    fetch_projects,
    fetch_project_detail,
    fetch_overview,
    parse_percent,
    parse_float,
    format_duration,
)


# ---------------------------------------------------------------------------
# Excel 工具函数
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="center")


def write_sheet(wb, sheet_name, headers, rows):
    """向 workbook 写入一个工作表。

    Args:
        wb: Workbook 实例
        sheet_name: 工作表名称
        headers: 表头列表，每项为 (key, label)
        rows: 数据行列表，每行为 dict
    """
    ws = wb.create_sheet(title=sheet_name)

    # 写表头
    for col, (_, label) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGNMENT

    # 写数据
    for row_idx, row_data in enumerate(rows, 2):
        for col, (key, _) in enumerate(headers, 1):
            val = row_data.get(key, "")
            # 避免写入 dict/list 类型（转为字符串）
            if isinstance(val, (dict, list)):
                val = str(val)
            ws.cell(row=row_idx, column=col, value=val)

    # 自动列宽
    for col, (_, label) in enumerate(headers, 1):
        max_len = len(str(label))
        for row_idx in range(2, min(ws.max_row + 1, 102)):  # 采样前 100 行
            cell_val = ws.cell(row=row_idx, column=col).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)

    return ws


# ---------------------------------------------------------------------------
# 各分析类型的导出逻辑
# ---------------------------------------------------------------------------

def export_top_consumers(headers, args, wb):
    """导出：任务资源消耗排行。"""
    tasks = fetch_all_tasks(
        headers,
        start_time=args.start_time,
        end_time=args.end_time,
        task_type=args.task_type,
    )

    scored = []
    for t in tasks:
        duration_ms = t.get("taskDuration", 0) or 0
        duration_h = duration_ms / (1000 * 3600)
        gpu_total = parse_float(t.get("gpu", {}).get("total"))
        cpu_total = parse_float(t.get("cpu", {}).get("total"))
        mem_total = parse_float(t.get("mem", {}).get("total"))
        gpu_hours = gpu_total * duration_h
        cpu_hours = cpu_total * duration_h
        composite = gpu_hours * 10 + cpu_hours + mem_total * 0.5

        scored.append({
            "taskName": t.get("taskName"),
            "userName": t.get("userName"),
            "taskType": t.get("taskType"),
            "taskStatus": t.get("taskStatus"),
            "startTime": t.get("startTime"),
            "duration": format_duration(duration_ms),
            "gpuTotal": gpu_total,
            "cpuTotal": cpu_total,
            "memTotal": mem_total,
            "gpuHours": round(gpu_hours, 2),
            "cpuHours": round(cpu_hours, 2),
            "compositeScore": round(composite, 2),
            "resourceGroupName": t.get("resourceGroupName", ""),
        })

    scored.sort(key=lambda x: x["compositeScore"], reverse=True)
    top = scored[:args.top]

    headers_spec = [
        ("taskName", "任务名称"),
        ("userName", "用户"),
        ("taskType", "任务类型"),
        ("taskStatus", "状态"),
        ("startTime", "开始时间"),
        ("duration", "运行时长"),
        ("gpuTotal", "GPU分配"),
        ("cpuTotal", "CPU分配"),
        ("memTotal", "内存分配"),
        ("gpuHours", "GPU·小时"),
        ("cpuHours", "CPU·小时"),
        ("compositeScore", "综合消耗分"),
        ("resourceGroupName", "资源组"),
    ]
    write_sheet(wb, "资源消耗排行", headers_spec, top)
    return len(top)


def export_user_resource_stats(headers, args, wb):
    """导出：用户资源统计。"""
    resources_gpu = fetch_all_resources(headers, resource_type="GPU")
    tasks = fetch_all_tasks(headers, start_time=args.start_time, end_time=args.end_time)

    user_stats = {}
    for t in tasks:
        user = t.get("userName", "unknown")
        if user not in user_stats:
            user_stats[user] = {"taskCount": 0, "runningTasks": 0,
                                "totalGpuHours": 0.0, "totalCpuHours": 0.0,
                                "totalDurationMs": 0}
        s = user_stats[user]
        s["taskCount"] += 1
        if t.get("taskStatus") == "Running":
            s["runningTasks"] += 1
        duration_h = (t.get("taskDuration", 0) or 0) / (1000 * 3600)
        s["totalGpuHours"] += parse_float(t.get("gpu", {}).get("total")) * duration_h
        s["totalCpuHours"] += parse_float(t.get("cpu", {}).get("total")) * duration_h
        s["totalDurationMs"] += t.get("taskDuration", 0) or 0

    rows = []
    for user, s in sorted(user_stats.items(), key=lambda x: x[1]["totalGpuHours"], reverse=True):
        gpu_alloc = next((r for r in resources_gpu if r.get("userName") == user), None)
        rows.append({
            "userName": user,
            "taskCount": s["taskCount"],
            "runningTasks": s["runningTasks"],
            "totalDuration": format_duration(s["totalDurationMs"]),
            "totalGpuHours": round(s["totalGpuHours"], 2),
            "totalCpuHours": round(s["totalCpuHours"], 2),
            "gpuAssigned": gpu_alloc.get("resourceAllocation", {}).get("assigned", "") if gpu_alloc else "",
            "gpuTotal": gpu_alloc.get("resourceAllocation", {}).get("total", "") if gpu_alloc else "",
            "gpuUsage": gpu_alloc.get("resourceUsage", "") if gpu_alloc else "",
        })

    headers_spec = [
        ("userName", "用户名"),
        ("taskCount", "任务总数"),
        ("runningTasks", "运行中"),
        ("totalDuration", "总运行时长"),
        ("totalGpuHours", "GPU·小时"),
        ("totalCpuHours", "CPU·小时"),
        ("gpuAssigned", "GPU已分配"),
        ("gpuTotal", "GPU总量"),
        ("gpuUsage", "GPU使用率"),
    ]
    write_sheet(wb, "用户资源统计", headers_spec, rows)
    return len(rows)


def export_gpu_split_stats(headers, args, wb):
    """导出：加速卡切分统计。"""
    cards, card_info = fetch_accel_cards(headers)
    specs = fetch_resource_specs(headers)

    # Sheet 1: 物理卡详情
    card_rows = []
    for c in cards:
        card_rows.append({
            "cardName": c.get("cardName"),
            "hostName": c.get("hostName"),
            "cardType": c.get("cardType"),
            "cardMode": c.get("cardMode"),
            "hadVirtualization": "是" if c.get("hadVirtualization") else "否",
            "utilization": c.get("utilization"),
            "memoryUtilization": c.get("memoryUtilization"),
            "temperature": c.get("temperature"),
            "power": c.get("power"),
            "taskNumber": c.get("taskNumber"),
        })
    write_sheet(wb, "物理卡详情", [
        ("cardName", "卡名"), ("hostName", "所在节点"), ("cardType", "型号"),
        ("cardMode", "模式"), ("hadVirtualization", "已虚拟化"),
        ("utilization", "GPU利用率"), ("memoryUtilization", "显存利用率"),
        ("temperature", "温度(°C)"), ("power", "功耗(W)"), ("taskNumber", "任务数"),
    ], card_rows)

    # Sheet 2: 资源规格
    gpu_specs = [s for s in specs if s.get("specType") == "gpu"]
    spec_rows = []
    for s in gpu_specs:
        spec_rows.append({
            "name": s.get("name"),
            "gpuType": s.get("gpuType"),
            "aiCore": s.get("aiCore"),
            "aiMemory": s.get("aiMemory"),
            "cpu": s.get("cpu"),
            "memory": s.get("memory"),
            "isFullCard": "整卡" if s.get("isFullCard") else "切分",
        })
    write_sheet(wb, "资源规格", [
        ("name", "规格名称"), ("gpuType", "GPU型号"), ("aiCore", "算力(%)"),
        ("aiMemory", "显存(G)"), ("cpu", "CPU(核)"), ("memory", "内存(G)"),
        ("isFullCard", "整卡/切分"),
    ], spec_rows)

    return len(card_rows)


def export_health_report(headers, args, wb):
    """导出：健康状态报表。"""
    clusters = fetch_clusters(headers)
    node_data = fetch_nodes(headers)
    cards, _ = fetch_accel_cards(headers)

    # Sheet 1: 集群
    cluster_rows = []
    for c in clusters:
        healthy = c.get("clusterStatus") == 1
        cluster_rows.append({
            "clusterName": c.get("clusterName"),
            "clusterStatus": "正常" if healthy else "异常",
            "version": c.get("version"),
            "serviceIp": c.get("serviceIp"),
            "healthy": "✓" if healthy else "✗",
        })
    write_sheet(wb, "集群状态", [
        ("clusterName", "集群名"), ("clusterStatus", "状态"),
        ("version", "版本"), ("serviceIp", "服务IP"), ("healthy", "健康"),
    ], cluster_rows)

    # Sheet 2: 节点
    node_rows = []
    for n in node_data["nodes"]:
        status = n.get("status", "")
        sched = n.get("schedulingAble", 0)
        healthy = status == "Ready" and sched == 1
        node_rows.append({
            "nodeName": n.get("nodeName"),
            "nodeIp": n.get("nodeIp"),
            "clusterName": n.get("clusterName"),
            "status": status,
            "schedulingAble": "可调度" if sched == 1 else "不可调度",
            "cpuLimit": n.get("cpuLimit"),
            "memLimit": n.get("memLimit"),
            "gpuNum": n.get("gpuNum"),
            "healthy": "✓" if healthy else "✗",
        })
    write_sheet(wb, "节点状态", [
        ("nodeName", "节点名"), ("nodeIp", "IP"), ("clusterName", "集群"),
        ("status", "状态"), ("schedulingAble", "调度"), ("cpuLimit", "CPU"),
        ("memLimit", "内存"), ("gpuNum", "GPU数"), ("healthy", "健康"),
    ], node_rows)

    # Sheet 3: 加速卡
    card_rows = []
    for c in cards:
        card_status = c.get("cardStatus", -1)
        temp = parse_float(c.get("temperature"))
        util = parse_percent(c.get("utilization"))
        task_num = c.get("taskNumber", 0)
        healthy = card_status == 0 and temp <= 85 and not (util < 1 and task_num > 0)

        reasons = []
        if card_status != 0:
            reasons.append(f"状态码异常({card_status})")
        if temp > 85:
            reasons.append(f"温度过高({temp}°C)")
        if util < 1 and task_num > 0:
            reasons.append("有任务但GPU利用率为0%")

        card_rows.append({
            "cardName": c.get("cardName"),
            "hostName": c.get("hostName"),
            "cardType": c.get("cardType"),
            "utilization": c.get("utilization"),
            "memoryUtilization": c.get("memoryUtilization"),
            "temperature": c.get("temperature"),
            "power": c.get("power"),
            "taskNumber": task_num,
            "healthy": "✓" if healthy else "✗",
            "abnormalReasons": "; ".join(reasons) if reasons else "",
        })
    write_sheet(wb, "加速卡状态", [
        ("cardName", "卡名"), ("hostName", "节点"), ("cardType", "型号"),
        ("utilization", "GPU利用率"), ("memoryUtilization", "显存利用率"),
        ("temperature", "温度(°C)"), ("power", "功耗(W)"), ("taskNumber", "任务数"),
        ("healthy", "健康"), ("abnormalReasons", "异常原因"),
    ], card_rows)

    return len(cluster_rows) + len(node_rows) + len(card_rows)


def export_task_status_dist(headers, args, wb):
    """导出：任务状态分布。"""
    tasks = fetch_all_tasks(headers)

    by_user = {}
    for t in tasks:
        user = t.get("userName", "unknown")
        status = t.get("taskStatus", "Unknown")
        if user not in by_user:
            by_user[user] = defaultdict(int)
        by_user[user][status] += 1

    # Sheet 1: 按用户统计
    user_rows = []
    for user, dist in sorted(by_user.items(), key=lambda x: sum(x[1].values()), reverse=True):
        row = {"userName": user, "total": sum(dist.values())}
        row.update(dict(dist))
        user_rows.append(row)

    # 动态表头：收集所有出现过的状态
    all_statuses = sorted(set(s for d in by_user.values() for s in d.keys()))
    headers_spec = [("userName", "用户名"), ("total", "总计")]
    for status in all_statuses:
        headers_spec.append((status, status))

    write_sheet(wb, "按用户统计", headers_spec, user_rows)

    # Sheet 2: 按类型统计
    by_type = {}
    for t in tasks:
        ttype = t.get("taskType", "unknown")
        status = t.get("taskStatus", "Unknown")
        if ttype not in by_type:
            by_type[ttype] = defaultdict(int)
        by_type[ttype][status] += 1

    type_rows = []
    for ttype, dist in sorted(by_type.items(), key=lambda x: sum(x[1].values()), reverse=True):
        row = {"taskType": ttype, "total": sum(dist.values())}
        row.update(dict(dist))
        type_rows.append(row)

    type_headers = [("taskType", "任务类型"), ("total", "总计")]
    for status in sorted(set(s for d in by_type.values() for s in d.keys())):
        type_headers.append((status, status))
    write_sheet(wb, "按类型统计", type_headers, type_rows)

    return len(user_rows) + len(type_rows)


def export_user_ranking(headers, args, wb):
    """导出：活跃度排行。"""
    tasks = fetch_all_tasks(headers)

    user_agg = {}
    for t in tasks:
        user = t.get("userName", "unknown")
        if user not in user_agg:
            user_agg[user] = {"taskCount": 0, "totalDurationMs": 0,
                              "totalGpuAllocated": 0.0, "totalCpuAllocated": 0.0}
        a = user_agg[user]
        a["taskCount"] += 1
        a["totalDurationMs"] += t.get("taskDuration", 0) or 0
        a["totalGpuAllocated"] += parse_float(t.get("gpu", {}).get("total"))
        a["totalCpuAllocated"] += parse_float(t.get("cpu", {}).get("total"))

    rows = []
    for user, a in user_agg.items():
        rows.append({
            "userName": user,
            "taskCount": a["taskCount"],
            "totalDuration": format_duration(a["totalDurationMs"]),
            "totalDurationMs": a["totalDurationMs"],
            "totalGpuAllocated": round(a["totalGpuAllocated"], 2),
            "totalCpuAllocated": round(a["totalCpuAllocated"], 2),
        })

    sort_key = args.sort_by
    if sort_key == "duration":
        rows.sort(key=lambda x: x["totalDurationMs"], reverse=True)
    elif sort_key == "gpu":
        rows.sort(key=lambda x: x["totalGpuAllocated"], reverse=True)
    elif sort_key == "cpu":
        rows.sort(key=lambda x: x["totalCpuAllocated"], reverse=True)
    else:
        rows.sort(key=lambda x: x["taskCount"], reverse=True)

    write_sheet(wb, "活跃度排行", [
        ("userName", "用户名"),
        ("taskCount", "任务总数"),
        ("totalDuration", "总运行时长"),
        ("totalGpuAllocated", "GPU总分配"),
        ("totalCpuAllocated", "CPU总分配"),
    ], rows)
    return len(rows)


def export_idle_detection(headers, args, wb):
    """导出：空闲资源检测。"""
    tasks = fetch_all_tasks(headers, task_status="Running")

    idle_threshold = args.idle_threshold
    duration_threshold_ms = args.duration_threshold * 3600 * 1000

    idle_tasks = []
    for t in tasks:
        duration_ms = t.get("taskDuration", 0) or 0
        gpu_usage = parse_percent(t.get("gpu", {}).get("usage"))
        cpu_usage = parse_percent(t.get("cpu", {}).get("usage"))
        mem_usage = parse_percent(t.get("mem", {}).get("usage"))

        all_low = gpu_usage < idle_threshold and cpu_usage < idle_threshold and mem_usage < idle_threshold
        long_running = duration_ms > duration_threshold_ms

        if all_low and long_running:
            idle_tasks.append({
                "taskName": t.get("taskName"),
                "userName": t.get("userName"),
                "taskType": t.get("taskType"),
                "duration": format_duration(duration_ms),
                "startTime": t.get("startTime"),
                "gpuUsage": t.get("gpu", {}).get("usage"),
                "cpuUsage": t.get("cpu", {}).get("usage"),
                "memUsage": t.get("mem", {}).get("usage"),
                "gpuTotal": parse_float(t.get("gpu", {}).get("total")),
                "cpuTotal": parse_float(t.get("cpu", {}).get("total")),
                "memTotal": parse_float(t.get("mem", {}).get("total")),
                "resourceGroupName": t.get("resourceGroupName", ""),
            })

    idle_tasks.sort(key=lambda x: x["durationMs"] if "durationMs" in x else 0, reverse=True)

    write_sheet(wb, "空闲任务检测", [
        ("taskName", "任务名称"), ("userName", "用户"), ("taskType", "类型"),
        ("duration", "运行时长"), ("startTime", "开始时间"),
        ("gpuUsage", "GPU使用率"), ("cpuUsage", "CPU使用率"), ("memUsage", "内存使用率"),
        ("gpuTotal", "GPU分配"), ("cpuTotal", "CPU分配"), ("memTotal", "内存分配"),
        ("resourceGroupName", "资源组"),
    ], idle_tasks)
    return len(idle_tasks)


def export_project_overview(headers, args, wb):
    """导出：项目资源概览。"""
    projects = fetch_projects(headers)

    # Sheet 1: 项目概览
    project_rows = []
    for p in projects:
        pid = p.get("projectId")
        try:
            detail = fetch_project_detail(headers, pid)
        except Exception:
            detail = {}

        specs = detail.get("resourceSpec", [])
        gpu_quota = [s for s in specs if s.get("specType") == "gpu"]
        quota_str = "; ".join(
            f"{s['name']}({s.get('useNum',0)}/{s.get('totalNum',0)})"
            for s in specs
        ) if specs else "无配额"

        members = detail.get("member", {}).get("list", [])
        member_str = ", ".join(m.get("name", "") for m in members) if members else "无成员"

        project_rows.append({
            "projectId": pid,
            "projectName": p.get("projectName"),
            "describes": p.get("describes") or detail.get("describes", ""),
            "creator": p.get("creator"),
            "memberNum": p.get("memberNum"),
            "members": member_str,
            "quota": quota_str,
            "createTime": p.get("createTime"),
        })

    write_sheet(wb, "项目概览", [
        ("projectId", "项目ID"), ("projectName", "项目名称"),
        ("describes", "描述"), ("creator", "创建者"),
        ("memberNum", "成员数"), ("members", "成员列表"),
        ("quota", "资源配额"), ("createTime", "创建时间"),
    ], project_rows)

    # Sheet 2: 配额明细
    quota_rows = []
    for p in projects:
        pid = p.get("projectId")
        pname = p.get("projectName")
        try:
            detail = fetch_project_detail(headers, pid)
        except Exception:
            continue
        for s in detail.get("resourceSpec", []):
            quota_rows.append({
                "projectName": pname,
                "specName": s.get("name"),
                "specType": s.get("specType"),
                "gpuType": s.get("gpuType", ""),
                "useNum": s.get("useNum", 0),
                "totalNum": s.get("totalNum", 0),
                "cpu": s.get("cpu"),
                "memory": s.get("memory"),
            })

    write_sheet(wb, "配额明细", [
        ("projectName", "项目"), ("specName", "规格名称"),
        ("specType", "类型"), ("gpuType", "GPU型号"),
        ("useNum", "已使用"), ("totalNum", "总量"),
        ("cpu", "CPU"), ("memory", "内存"),
    ], quota_rows)

    return len(project_rows)


# ---------------------------------------------------------------------------
# 导出函数映射
# ---------------------------------------------------------------------------

EXPORT_HANDLERS = {
    "top-consumers": export_top_consumers,
    "user-resource-stats": export_user_resource_stats,
    "gpu-split-stats": export_gpu_split_stats,
    "health-report": export_health_report,
    "task-status-dist": export_task_status_dist,
    "user-ranking": export_user_ranking,
    "idle-detection": export_idle_detection,
    "project-overview": export_project_overview,
}


def main():
    parser = argparse.ArgumentParser(description="报表导出为 Excel")
    parser.add_argument("--analysis", required=True, choices=EXPORT_HANDLERS.keys(),
                        help="分析类型")
    parser.add_argument("--output", required=True, help="输出文件路径（.xlsx）")
    # 通用筛选参数（与 analyze_report 一致）
    parser.add_argument("--start-time", default=None)
    parser.add_argument("--end-time", default=None)
    parser.add_argument("--cluster-name", default=None)
    parser.add_argument("--task-type", default=None)
    parser.add_argument("--user-name", default=None)
    parser.add_argument("--top", type=int, default=10)
    parser.add_argument("--sort-by", default="taskCount",
                        choices=["taskCount", "duration", "gpu", "cpu"])
    parser.add_argument("--idle-threshold", type=float, default=5.0)
    parser.add_argument("--duration-threshold", type=float, default=1.0)
    args = parser.parse_args()

    if not args.output.endswith(".xlsx"):
        args.output += ".xlsx"

    headers = make_admin_headers({"Content-Type": "application/json;charset=UTF-8"})

    wb = Workbook()
    # 删除默认的空 Sheet
    wb.remove(wb.active)

    handler = EXPORT_HANDLERS[args.analysis]
    row_count = handler(headers, args, wb)

    wb.save(args.output)

    abs_path = os.path.abspath(args.output)
    output_result(
        success=True,
        message=f"已导出 {args.analysis} 报表到 {abs_path}（{row_count} 行数据）",
        filePath=abs_path,
        analysisType=args.analysis,
        rowCount=row_count,
    )


if __name__ == "__main__":
    main()
